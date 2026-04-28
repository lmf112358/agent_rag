"""
Excel解析器

负责读取报价清单Excel并映射为标准字段。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - 依赖缺失时在运行期报错
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover - 依赖缺失时在运行期报错
    xlrd = None

from .config import (
    COLUMN_MAPPING,
    PHASE_PATTERNS,
    SECTION_SUBTOTAL_PATTERNS,
    SUMMARY_ROW_PATTERNS,
)
from .models import ExcelSheetResult


class ExcelParser:
    """报价清单Excel解析器（优先openpyxl）"""

    def parse(self, excel_path: str) -> List[ExcelSheetResult]:
        """解析Excel并返回每个Sheet的标准化原始结果。"""
        path = Path(excel_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")

        suffix = path.suffix.lower()
        if suffix == ".xls":
            return self._parse_xls(excel_path)
        elif suffix == ".xlsx":
            return self._parse_xlsx(excel_path)
        else:
            raise ValueError(f"不支持的Excel格式: {suffix}，仅支持 .xls 和 .xlsx")

    def _parse_xlsx(self, excel_path: str) -> List[ExcelSheetResult]:
        """使用 openpyxl 解析 .xlsx 文件"""
        if load_workbook is None:
            raise ImportError("缺少依赖 openpyxl，请先安装后再执行报价审核")

        workbook = load_workbook(filename=excel_path, data_only=True)
        return self._parse_workbook(workbook)

    def _parse_xls(self, excel_path: str) -> List[ExcelSheetResult]:
        """使用 xlrd 解析 .xls 文件"""
        if xlrd is None:
            raise ImportError("缺少依赖 xlrd，请先安装：pip install xlrd")

        workbook = xlrd.open_workbook(excel_path)
        results: List[ExcelSheetResult] = []

        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            sheet_name = sheet.name

            rows: List[Tuple[Any, ...]] = []
            for row_idx in range(sheet.nrows):
                rows.append(tuple(sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)))

            if not rows:
                continue

            header_row_index, header_map, raw_headers = self._detect_header(rows)
            if header_row_index is None or not header_map:
                continue

            parsed_rows: List[Dict[str, Any]] = []
            section_breaks: List[int] = []

            for row_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
                if self._is_empty_row(row):
                    continue

                row_dict = self._map_row(row, header_map)
                row_dict["_row_index"] = row_idx

                first_cell = self._safe_text(row[0] if row else "")
                if self._is_section_break_row(first_cell):
                    section_breaks.append(row_idx)

                parsed_rows.append(row_dict)

            results.append(
                ExcelSheetResult(
                    sheet_name=sheet_name,
                    phase=self._detect_phase(sheet_name),
                    raw_headers=raw_headers,
                    rows=parsed_rows,
                    section_breaks=section_breaks,
                )
            )

        return results

    def _parse_workbook(self, workbook) -> List[ExcelSheetResult]:
        """通用工作簿解析逻辑（供 openpyxl 使用）"""
        results: List[ExcelSheetResult] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_row_index, header_map, raw_headers = self._detect_header(rows)
            if header_row_index is None or not header_map:
                continue

            parsed_rows: List[Dict[str, Any]] = []
            section_breaks: List[int] = []

            for row_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
                if self._is_empty_row(row):
                    continue

                row_dict = self._map_row(row, header_map)
                row_dict["_row_index"] = row_idx

                first_cell = self._safe_text(row[0] if row else "")
                if self._is_section_break_row(first_cell):
                    section_breaks.append(row_idx)

                parsed_rows.append(row_dict)

            results.append(
                ExcelSheetResult(
                    sheet_name=sheet_name,
                    phase=self._detect_phase(sheet_name),
                    raw_headers=raw_headers,
                    rows=parsed_rows,
                    section_breaks=section_breaks,
                )
            )

        return results

    def _detect_header(self, rows: List[Tuple[Any, ...]]) -> Tuple[Optional[int], Dict[str, int], List[str]]:
        best_idx: Optional[int] = None
        best_map: Dict[str, int] = {}
        best_headers: List[str] = []

        for idx, row in enumerate(rows[:30]):
            headers = [self._safe_text(value) for value in row]
            mapped = self._build_header_mapping(headers)
            if len(mapped) > len(best_map):
                best_idx = idx
                best_map = mapped
                best_headers = headers

        if len(best_map) < 4:
            return None, {}, []

        return best_idx, best_map, best_headers

    def _build_header_mapping(self, headers: List[str]) -> Dict[str, int]:
        normalized_headers = [self._normalize_header(h) for h in headers]
        mapping: Dict[str, int] = {}

        for canonical, aliases in COLUMN_MAPPING.items():
            alias_set = {self._normalize_header(alias) for alias in aliases}
            for i, h in enumerate(normalized_headers):
                if h in alias_set:
                    mapping[canonical] = i
                    break

        return mapping

    def _map_row(self, row: Tuple[Any, ...], header_map: Dict[str, int]) -> Dict[str, Any]:
        row_dict: Dict[str, Any] = {}
        for field, index in header_map.items():
            value = row[index] if index < len(row) else None
            row_dict[field] = value
        return row_dict

    def _detect_phase(self, sheet_name: str) -> str:
        text = self._safe_text(sheet_name)
        for phase, keywords in PHASE_PATTERNS.items():
            if any(keyword.lower() in text.lower() for keyword in keywords):
                return phase
        return "未知"

    def _is_section_break_row(self, text: str) -> bool:
        if not text:
            return False

        if any(keyword in text for keyword in SECTION_SUBTOTAL_PATTERNS):
            return True

        for keywords in SUMMARY_ROW_PATTERNS.values():
            if any(keyword in text for keyword in keywords):
                return True

        return False

    def _is_empty_row(self, row: Tuple[Any, ...]) -> bool:
        return all(self._safe_text(value) == "" for value in row)

    def _safe_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize_header(self, header: str) -> str:
        return "".join(self._safe_text(header).lower().split())
